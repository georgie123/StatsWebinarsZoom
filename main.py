import os
from datetime import date
from tabulate import tabulate as tab

import pandas as pd

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
from matplotlib.collections import PatchCollection
from mpl_toolkits.basemap import Basemap

import numpy as np

from PIL import Image, ImageOps

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

today = date.today()

shp_simple_countries = r'C:/Users/Georges/PycharmProjects/data/simple_countries/simple_countries'
shp_simple_areas = r'C:/Users/Georges/PycharmProjects/data/simple_areas/simple_areas'
inputCountryConversion = r'C:/Users/Georges/PycharmProjects/data/countries_conversion.xlsx'

workDirectory = r'C:/Users/Georges/Downloads/Webinar/'

WebinarFileName = '20210225_Webinar_HpvAssays_Arbyn'

ReportExcelFile = workDirectory + WebinarFileName + '_Report.xlsx'
NewAddThenDeleteExcelFile = workDirectory + WebinarFileName + '_NewAddJooThenDelete.xlsx'
NewCollectExcelFile = workDirectory + WebinarFileName + '_NewToCollect.xlsx'


# WEBINAR EXCEL IMPORT
inputExcelFile = workDirectory+WebinarFileName+'.csv'
df = pd.read_csv(inputExcelFile, sep=',',
                   usecols=['Attended', 'First Name', 'Last Name', 'Email', 'Phone', 'Specialty',
                            'City', 'Country/Region Name'
                            ])


# REMOVE ADMIN AND TEAM
index_drop1 = df[df['Email'].apply(lambda x: x.endswith('@informa.com'))].index
df = df.drop(index_drop1)
index_drop2 = df[df['Email'].apply(lambda x: x.endswith('@euromedicom.com'))].index
df = df.drop(index_drop2)
index_drop3 = df[df['Email'].apply(lambda x: x.endswith('@eurogin.com'))].index
df = df.drop(index_drop3)
index_drop4 = df[df['Email'].apply(lambda x: x.endswith('@multispecialtysociety.com'))].index
df = df.drop(index_drop4)
index_drop5 = df[df['Email'].apply(lambda x: x.endswith('@ce.com.co'))].index
df = df.drop(index_drop5)
index_drop6 = df[df['Email'].apply(lambda x: x == ('max.carter11@yahoo.com'))].index
df = df.drop(index_drop6)


# ATTENDED YES, BEFORE DEDUPE
dfAttendedYes = df[df['Attended'] == 'Yes']

# ATTENDED NO, BEFORE DEDUPE
dfAttendedNo = df[df['Attended'] == 'No']

dfCumulated = pd.concat([dfAttendedYes, dfAttendedNo], ignore_index=True)


# DEDUPE
dfAllDeduped = dfCumulated.drop_duplicates(subset=['Email'], keep='first')
dfAttendedDeduped = dfAttendedYes.drop_duplicates(subset=['Email'], keep='first')



# EUROGIN_ACYMAILING_SUBSCRIBER IMPORT
df_subscriber = pd.read_csv(workDirectory+'eurogin_acymailing_subscriber.csv', sep=',', quotechar='"', usecols=['source', 'email'])
# SOURCES HARMONIZATION
df_subscriber['source'] = df_subscriber['source'].replace({'EXTERN: ': ''}, regex=True)
df_subscriber['source'] = df_subscriber['source'].replace({'PROSPECT: ': ''}, regex=True)


# COUNTRY CONVERSION IMPORT
df_CountryConversion = pd.read_excel(inputCountryConversion, sheet_name='countries', engine='openpyxl',
                   usecols=['CTRY_ZOOM', 'continent_stat'])


# NEW EMAILS
df_WebinarNew = pd.DataFrame(dfAllDeduped[~dfAllDeduped['Email'].str.lower().isin(df_subscriber['email'].str.lower())])
newWebinar = df_WebinarNew.shape[0]


# COUNT PARTICIPANTS: YES, NO, ALL REGISTERED
attended = dfAllDeduped[dfAllDeduped['Attended'] == 'Yes'].count()['Email']
noshow = dfAllDeduped[dfAllDeduped['Attended'] == 'No'].count()['Email']
registered = dfAllDeduped.shape[0]

participantLabel = []
myCounts = []
participantLabel.extend(('Attended', 'No-show', 'Total'))
myCounts.extend((attended, noshow, registered))

participantDict = list(zip(participantLabel, myCounts))
df_ParticipantCount = pd.DataFrame(participantDict, columns=['Participants', 'Total'])

df_ParticipantCount['%'] = (df_ParticipantCount['Total'] / registered) * 100
df_ParticipantCount['%'] = df_ParticipantCount['%'].round(decimals=2)


# COUNT COUNTRY
df_Country_count = pd.DataFrame(dfAttendedDeduped.groupby(['Country/Region Name'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Country_count = df_Country_count.fillna('Unknow')

df_Country_count['Percent'] = (df_Country_count['Total'] / df_Country_count['Total'].sum()) * 100
df_Country_count['Percent'] = df_Country_count['Percent'].round(decimals=1)


# COUNT AREAS
# JOIN LEFT WITH COUNTRY CONVERSION
df_WebinarAreas = pd.merge(dfAttendedDeduped, df_CountryConversion, left_on='Country/Region Name', right_on='CTRY_ZOOM', how='left')\
    [['Email', 'continent_stat']]

df_AreasCount = pd.DataFrame(df_WebinarAreas.groupby(['continent_stat'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_AreasCount = df_AreasCount.fillna('Unknow')

df_AreasCount['Percent'] = (df_AreasCount['Total'] / df_AreasCount['Total'].sum()) * 100
df_AreasCount['Percent'] = df_AreasCount['Percent'].round(decimals=1)


# COUNT SPECIALTIES
df_Specialties_count = pd.DataFrame(dfAttendedDeduped.groupby(['Specialty'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Specialties_count = df_Specialties_count.fillna('Unknow')

df_Specialties_count['Percent'] = (df_Specialties_count['Total'] / df_Specialties_count['Total'].sum()) * 100
df_Specialties_count['Percent'] = df_Specialties_count['Percent'].round(decimals=1)


# COUNT SPECIALTIES PER COUNTRY
df_SpecialtiesPerCountry_count = pd.DataFrame(dfAttendedDeduped.groupby(['Country/Region Name', 'Specialty'], dropna=False)\
    .size(), columns=['Total']).sort_values(['Country/Region Name', 'Total'], ascending=[True, False]).reset_index()
df_SpecialtiesPerCountry_count = df_SpecialtiesPerCountry_count.fillna('Unknow')

df_SpecialtiesPerCountry_count['Percent'] = (df_SpecialtiesPerCountry_count['Total'] / df_SpecialtiesPerCountry_count['Total'].sum()) * 100
df_SpecialtiesPerCountry_count['Percent'] = df_SpecialtiesPerCountry_count['Percent'].round(decimals=2)


# COUNT SOURCES
# JOIN LEFT WITH SUBSCRIBERS
df_WebinarSubscriber = pd.merge(dfAttendedDeduped, df_subscriber, left_on='Email', right_on='email', how='left')\
    [['Email', 'source']]

df_Sources = pd.DataFrame(df_WebinarSubscriber.groupby(['source'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Sources = df_Sources.fillna('')

df_Sources['Percent'] = (df_Sources['Total'] / df_Sources['Total'].sum()) * 100
df_Sources['Percent'] = df_Sources['Percent'].round(decimals=1)


# COUNT NEW COUNTRY
df_NewCountry_count = pd.DataFrame(df_WebinarNew.groupby(['Country/Region Name'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewCountry_count = df_NewCountry_count.fillna('Unknow')

df_NewCountry_count['Percent'] = (df_NewCountry_count['Total'] / df_NewCountry_count['Total'].sum()) * 100
df_NewCountry_count['Percent'] = df_NewCountry_count['Percent'].round(decimals=1)


# COUNT NEW AREAS
# JOIN LEFT WITH COUNTRY CONVERSION
df_NewWebinarAreas = pd.merge(df_WebinarNew, df_CountryConversion, left_on='Country/Region Name', right_on='CTRY_ZOOM', how='left')\
    [['Email', 'continent_stat']]

df_NewAreasCount = pd.DataFrame(df_NewWebinarAreas.groupby(['continent_stat'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewAreasCount = df_NewAreasCount.fillna('Unknow')

df_NewAreasCount['Percent'] = (df_NewAreasCount['Total'] / df_NewAreasCount['Total'].sum()) * 100
df_NewAreasCount['Percent'] = df_NewAreasCount['Percent'].round(decimals=1)


# COUNT NEW SPECIALTIES
df_NewSpecialties_count = pd.DataFrame(df_WebinarNew.groupby(['Specialty'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewSpecialties_count = df_NewSpecialties_count.fillna('Unknow')

df_NewSpecialties_count['Percent'] = (df_NewSpecialties_count['Total'] / df_NewSpecialties_count['Total'].sum()) * 100
df_NewSpecialties_count['Percent'] = df_NewSpecialties_count['Percent'].round(decimals=1)


# COUNT NEW SPECIALTIES PER COUNTRY
df_NewSpecialtiesPerCountry_count = pd.DataFrame(df_WebinarNew.groupby(['Country/Region Name', 'Specialty'], dropna=False)\
    .size(), columns=['Total']).sort_values(['Country/Region Name', 'Total'], ascending=[True, False]).reset_index()
df_NewSpecialtiesPerCountry_count = df_NewSpecialtiesPerCountry_count.fillna('Unknow')

df_NewSpecialtiesPerCountry_count['Percent'] = (df_NewSpecialtiesPerCountry_count['Total'] / df_NewSpecialtiesPerCountry_count['Total'].sum()) * 100
df_NewSpecialtiesPerCountry_count['Percent'] = df_NewSpecialtiesPerCountry_count['Percent'].round(decimals=2)


# EXCEL FILE: NEW TO ADD IN EUROGIN_ACYMAILING_SUBSCRIBER THEN DELETE
df_WebinarNew['source'] = WebinarFileName.replace('_', ' ').upper()
writer = pd.ExcelWriter(NewAddThenDeleteExcelFile, engine='xlsxwriter')
df_WebinarNew[['Email', 'source']].to_excel(writer, index=False, sheet_name='New Add Then Delete')
writer.save()


# EXCEL FILE: NEW TO COLLECT
writer = pd.ExcelWriter(NewCollectExcelFile, engine='xlsxwriter')
df_WebinarNew.to_excel(writer, index=False, sheet_name='New Collect')
writer.save()


# EXCEL FILE: REPORT
writer = pd.ExcelWriter(ReportExcelFile, engine='xlsxwriter')

df_ParticipantCount.to_excel(writer, index=False, sheet_name='Attendance', header=['Participants', 'Total', '%'])

df_Country_count.to_excel(writer, index=False, sheet_name='Countries attended', header=['Country', 'Total', '%'])
df_AreasCount.to_excel(writer, index=False, sheet_name='Areas attended', header=['Area', 'Total', '%'])
df_Specialties_count.to_excel(writer, index=False, sheet_name='Specialties attended', header=['Specialty', 'Total', '%'])
df_SpecialtiesPerCountry_count.to_excel(writer, index=False, sheet_name='SpecialtiesCountry attended', header=['Country', 'Specialty', 'Total', '%'])

df_Sources.to_excel(writer, index=False, sheet_name='Sources attended', header=['Source', 'Total', '%'])

df_NewCountry_count.to_excel(writer, index=False, sheet_name='News Countries', header=['Country', 'Total', '%'])
df_NewAreasCount.to_excel(writer, index=False, sheet_name='News Areas', header=['Area', 'Total', '%'])
df_NewSpecialties_count.to_excel(writer, index=False, sheet_name='News Specialties', header=['Specialty', 'Total', '%'])
df_NewSpecialtiesPerCountry_count.to_excel(writer, index=False, sheet_name='News Specialties per country', header=['Country', 'Specialty', 'Total', '%'])

writer.save()

# EXCEL FILTERS
workbook = openpyxl.load_workbook(ReportExcelFile)
sheetsLits = workbook.sheetnames

for sheet in sheetsLits:
    if sheet == 'Attendance':
        continue
    worksheet = workbook[sheet]
    FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    workbook.save(ReportExcelFile)

# EXCEL COLORS
for sheet in sheetsLits:
    worksheet = workbook[sheet]
    for cell in workbook[sheet][1]:
        worksheet[cell.coordinate].fill = PatternFill(fgColor = 'FFC6C1C1', fill_type = 'solid')
        workbook.save(ReportExcelFile)

# EXCEL COLUMN SIZE
for sheet in sheetsLits:
    for cell in workbook[sheet][1]:
        if get_column_letter(cell.column) == 'A':
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 30
        else:
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 10
        workbook.save(ReportExcelFile)


# WRITE NEW EMAILS
workbook['Attendance']['A7'].fill = PatternFill(fgColor = 'FFC6C1C1', fill_type = 'solid')
workbook['Attendance']['A7'] = 'New emails: '+str(newWebinar)
workbook['Attendance']['A7'].font = Font(bold=True)
workbook.save(ReportExcelFile)


# CHART PARTICIPANTS STATUS
participantLabel.pop()

participantValue = []
participantValue.extend([attended, noshow])

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig1 = plt.figure(figsize=(8, 5))
plt.pie(participantValue, labels=participantLabel, colors=colors, explode=[0.02, 0.02], autopct='%1.2f%%', shadow=False, startangle=90)
plt.axis('equal')
plt.title('Attendance', pad=20, fontsize=15)

# Footer
plt.figtext(0.2, 0.10, WebinarFileName.replace('_', ' '), ha="left", fontsize=13, weight='bold')
plt.figtext(0.2, 0.06, 'No-show: '+str(noshow)+' - New emails: '+str(newWebinar), ha="left", fontsize=11)

fig1.savefig(workDirectory+'myplot1.png', dpi=90)
plt.clf()

im = Image.open(workDirectory+'myplot1.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot1.png')

# INSERT CHART IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot1.png')
img.anchor = 'E2'

workbook['Attendance'].add_image(img)
workbook.save(ReportExcelFile)


# MAP COUNTRIES
df_Country_count.set_index('Country/Region Name', inplace=True)

my_values = df_Country_count['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_Country_count['Percent'] = np.digitize(my_values, my_range) - 1

map1 = plt.figure(figsize=(14, 8))

ax = map1.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_countries, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['CTRY_ZOOM']
    if shp_ctry not in df_Country_count.index:
        color = '#dddddd'
    else:
        color = scheme[df_Country_count.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map1.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' '), ha="left", fontsize=13, weight='bold')
plt.figtext(0.2, 0.14, 'No-show: '+str(noshow)+' - New emails: '+str(newWebinar), ha="left", fontsize=11)

cb.remove()

map1.savefig(workDirectory+'mymap1.png', dpi=110, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap1.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap1.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap1.png')
img.anchor = 'E2'

workbook['Countries attended'].add_image(img)
workbook.save(ReportExcelFile)


# CHART AREAS
chartLabel = df_AreasCount['continent_stat'].tolist()
chartLegendLabel = df_AreasCount['continent_stat'].tolist()
chartValue = df_AreasCount['Total'].tolist()
chartLegendPercent = df_AreasCount['Percent'].tolist()

chartLabel[-1] = ''

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig2 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Areas', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig2.savefig(workDirectory+'myplot2.png', dpi=80)
plt.clf()

im = Image.open(workDirectory+'myplot2.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot2.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot2.png')
img.anchor = 'A13'

workbook['Areas attended'].add_image(img)
workbook.save(ReportExcelFile)


# MAP AREAS
df_AreasCount.set_index('continent_stat', inplace=True)

my_values = df_AreasCount['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_AreasCount['Percent'] = np.digitize(my_values, my_range) - 1

map2 = plt.figure(figsize=(14, 8))

ax = map2.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_areas, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['continent']
    if shp_ctry not in df_AreasCount.index:
        color = '#dddddd'
    else:
        color = scheme[df_AreasCount.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map2.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' '), ha="left", fontsize=13, weight='bold')
plt.figtext(0.2, 0.14, 'No-show: '+str(noshow)+' - New emails: '+str(newWebinar), ha="left", fontsize=11)

cb.remove()

map2.savefig(workDirectory+'mymap2.png', dpi=90, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap2.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap2.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap2.png')
img.anchor = 'G2'

workbook['Areas attended'].add_image(img)
workbook.save(ReportExcelFile)


# MAP NEW COUNTRIES
df_NewCountry_count.set_index('Country/Region Name', inplace=True)

my_values = df_NewCountry_count['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_NewCountry_count['Percent'] = np.digitize(my_values, my_range) - 1

map3 = plt.figure(figsize=(14, 8))

ax = map3.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_countries, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['CTRY_ZOOM']
    if shp_ctry not in df_NewCountry_count.index:
        color = '#dddddd'
    else:
        color = scheme[df_NewCountry_count.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map3.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' '), ha="left", fontsize=13, weight='bold')

cb.remove()

map3.savefig(workDirectory+'mymap3.png', dpi=110, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap3.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap3.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap3.png')
img.anchor = 'E2'

workbook['News Countries'].add_image(img)
workbook.save(ReportExcelFile)


# CHART NEW AREAS
chartLabel = df_NewAreasCount['continent_stat'].tolist()
chartLegendLabel = df_NewAreasCount['continent_stat'].tolist()
chartValue = df_NewAreasCount['Total'].tolist()
chartLegendPercent = df_NewAreasCount['Percent'].tolist()

chartLabel[-1] = ''

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig3 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Areas', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig3.savefig(workDirectory+'myplot3.png', dpi=80)
plt.clf()

im = Image.open(workDirectory+'myplot3.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot3.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot3.png')
img.anchor = 'A13'

workbook['News Areas'].add_image(img)
workbook.save(ReportExcelFile)


# MAP NEW AREAS
df_NewAreasCount.set_index('continent_stat', inplace=True)

my_values = df_NewAreasCount['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_NewAreasCount['Percent'] = np.digitize(my_values, my_range) - 1

map4 = plt.figure(figsize=(14, 8))

ax = map4.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_areas, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['continent']
    if shp_ctry not in df_NewAreasCount.index:
        color = '#dddddd'
    else:
        color = scheme[df_NewAreasCount.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map4.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' '), ha="left", fontsize=13, weight='bold')

cb.remove()

map4.savefig(workDirectory+'mymap4.png', dpi=90, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap4.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap4.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap4.png')
img.anchor = 'G2'

workbook['News Areas'].add_image(img)
workbook.save(ReportExcelFile)


# REMOVE PICTURES
os.remove(workDirectory+'myplot1.png')
os.remove(workDirectory+'myplot2.png')
os.remove(workDirectory+'myplot3.png')
os.remove(workDirectory+'mymap1.png')
os.remove(workDirectory+'mymap2.png')
os.remove(workDirectory+'mymap3.png')
os.remove(workDirectory+'mymap4.png')


print(tab(df_AreasCount, headers='keys', tablefmt='psql', showindex=False))
print(newWebinar)